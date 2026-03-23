#!/usr/bin/env python3
"""
DAILY Excel 파일 → PostgreSQL 8개 테이블 일괄 삽입 스크립트
사용법: python3 scripts/bulk_insert_daily.py market/260304_DAILY.xlsx
"""

import sys
import os
import re
from datetime import datetime, date
from decimal import Decimal
import openpyxl
import psycopg2

DB_URL = "postgresql://neondb_owner:npg_IcbygUXf0Sa4@ep-sweet-smoke-adz6o5mz-pooler.c-2.us-east-1.aws.neon.tech/neondb?sslmode=require"


def cell(ws, r, c):
    """셀 값 반환 (None 안전)"""
    v = ws.cell(r, c).value
    return v


def num(v):
    """숫자 변환 (None/빈값 → None)"""
    if v is None or v == '' or v == '-' or v == 0:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def extract_base_date(ws):
    """파일에서 기준일 추출 (R3:AA3 또는 R4:AH4 부근 날짜)"""
    for r in range(3, 6):
        for c in range(25, 40):
            v = cell(ws, r, c)
            if isinstance(v, datetime):
                return v.date()
    # 파일명에서 추출 시도
    return None


def parse_date_from_filename(filepath):
    """파일명에서 날짜 추출: '260304_DAILY' 또는 '260323 데일리' → 2026-03-04"""
    basename = os.path.basename(filepath)
    # 패턴1: YYMMDD_DAILY (기존)
    m = re.search(r'(\d{6})_DAILY', basename)
    if not m:
        # 패턴2: YYMMDD 데일리 (신규 — 공백/언더스코어 + 한글)
        m = re.search(r'(\d{6})[\s_]*데일리', basename)
    if m:
        d = m.group(1)
        year = 2000 + int(d[:2])
        month = int(d[2:4])
        day = int(d[4:6])
        return date(year, month, day)
    return None


def insert_macro_index(ws, cur, base_date):
    """R9~R16: 주요 증시 및 환율"""
    rows_map = [
        (9, 'EQUITY', 'DOW'),
        (10, 'EQUITY', 'NASDAQ'),
        (11, 'EQUITY', 'S&P500'),
        (12, 'EQUITY', 'NIKKEI'),
        (13, 'EQUITY', 'KOSPI'),
        (14, 'EQUITY', 'KOSDAQ'),
        (15, 'FX', 'USD/KRW'),
        (16, 'FX', 'JPY/USD'),
    ]
    count = 0
    for row, asset_class, ticker in rows_map:
        close_val = num(cell(ws, row, 4))
        change_val = num(cell(ws, row, 5))
        change_pct = num(cell(ws, row, 7))
        if close_val is not None:
            cur.execute(
                "INSERT INTO tb_macro_index (base_date, asset_class, ticker, close_value, change_val, change_pct) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, asset_class, ticker, close_val, change_val, change_pct)
            )
            count += 1
    print(f"  tb_macro_index: {count}건")
    return count


def insert_domestic_rate(ws, cur, base_date):
    """IRS, CRS, CD/통안/회사채/국고채 금리"""
    count = 0

    # US Treasury: R9~R12 (col 9=ticker, 10=level, 11=change)
    for row in range(9, 13):
        tenor = cell(ws, row, 9)
        level = num(cell(ws, row, 10))
        change = num(cell(ws, row, 11))
        if tenor and level is not None:
            cur.execute(
                "INSERT INTO tb_domestic_rate (base_date, rate_type, maturity, ticker_name, yield_val, change_bp) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, 'UST', str(tenor), f'UST {tenor}', level, change)
            )
            count += 1

    # IRS: R9~ (col 15=만기, 16=IRS, 17=change)
    for row in range(9, 14):
        tenor = cell(ws, row, 15)
        irs_val = num(cell(ws, row, 16))
        change = num(cell(ws, row, 17))
        if tenor and irs_val is not None:
            cur.execute(
                "INSERT INTO tb_domestic_rate (base_date, rate_type, maturity, ticker_name, yield_val, change_bp) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, 'IRS', str(tenor), f'IRS {tenor}', irs_val, change)
            )
            count += 1

    # CRS: R9~ (col 19=CRS, 20=change)
    for row in range(9, 14):
        tenor = cell(ws, row, 15)
        crs_val = num(cell(ws, row, 19))
        change = num(cell(ws, row, 20))
        if tenor and crs_val is not None:
            cur.execute(
                "INSERT INTO tb_domestic_rate (base_date, rate_type, maturity, ticker_name, yield_val, change_bp) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, 'CRS', str(tenor), f'CRS {tenor}', crs_val, change)
            )
            count += 1

    # CD/통안/회사채/국고채: R14~R17 (col 22=구분, 24=level, 25=change)
    for row in range(14, 18):
        name = cell(ws, row, 22)
        level = num(cell(ws, row, 24))
        change = num(cell(ws, row, 25))
        if name and level is not None and str(name).strip() != '0':
            name_str = str(name).strip()
            # 분류
            if 'CD' in name_str:
                rate_type = 'CD'
            elif '통안' in name_str:
                rate_type = '통안'
            elif '회' in name_str:
                rate_type = '회사채'
            elif '국고' in name_str:
                rate_type = '국고채'
            else:
                rate_type = '기타'
            cur.execute(
                "INSERT INTO tb_domestic_rate (base_date, rate_type, maturity, ticker_name, yield_val, change_bp) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, rate_type, name_str, name_str, level, change)
            )
            count += 1

    print(f"  tb_domestic_rate: {count}건")
    return count


def insert_ktb_futures(ws, cur, base_date):
    """KTB 선물: R20~R23 (col 22~34)"""
    count = 0
    # KTB3: R20, KTB10: R22
    futures_rows = [
        (20, 'KTB 3Y'),
        (22, 'KTB 10Y'),
    ]
    for row, ticker in futures_rows:
        # 선물 순매수 데이터
        foreign = num(cell(ws, row, 23))
        individual = num(cell(ws, row, 25))
        fin_invest = num(cell(ws, row, 28))
        trust = num(cell(ws, row, 30))
        bank = num(cell(ws, row, 32))

        # 선물가 (R8:col31=종가)
        close_price = None
        volume = 0
        if ticker == 'KTB 3Y':
            close_price = num(cell(ws, 11, 31))  # R11:AE = 종가
            vol = num(cell(ws, 15, 31))
            volume = int(vol) if vol else 0
        elif ticker == 'KTB 10Y':
            close_price = num(cell(ws, 11, 34))
            vol = num(cell(ws, 15, 34))
            volume = int(vol) if vol else 0

        if foreign is not None:
            cur.execute(
                "INSERT INTO tb_ktb_futures (base_date, ticker, close_price, volume, open_interest, net_foreign, net_fin_invest, net_bank) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, ticker, close_price, volume, 0,
                 int(foreign) if foreign else 0,
                 int(fin_invest) if fin_invest else 0,
                 int(bank) if bank else 0)
            )
            count += 1

    print(f"  tb_ktb_futures: {count}건")
    return count


def insert_primary_market(ws, cur, base_date):
    """전일발행시장: R21~ (col 2=종목명, 5=등급, 7=만기, 9=금리, 11=수량, 16~19=spread)"""
    count = 0
    for row in range(21, 28):
        issuer = cell(ws, row, 2)
        if not issuer or str(issuer).strip() == '0' or str(issuer).strip() == '':
            continue
        issuer_str = str(issuer).strip()
        rating = str(cell(ws, row, 5) or '').strip()
        if rating == '0':
            rating = ''
        maturity = str(cell(ws, row, 7) or '').strip()
        if maturity == '0':
            maturity = ''
        issue_yield = num(cell(ws, row, 9))
        issue_amt_raw = cell(ws, row, 11)
        issue_amt = 0
        if issue_amt_raw:
            amt_str = str(issue_amt_raw).replace(',', '').replace('억', '').strip()
            try:
                issue_amt = int(float(amt_str))
            except:
                issue_amt = 0

        spread = num(cell(ws, row, 19))

        if issue_yield is not None:
            cur.execute(
                "INSERT INTO tb_primary_market (base_date, issuer_name, credit_rating, maturity, issue_yield, issue_amt, spread_bp) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, issuer_str, rating, maturity, issue_yield, issue_amt, spread)
            )
            count += 1

    print(f"  tb_primary_market: {count}건")
    return count


def insert_bond_valuation(ws, cur, base_date):
    """주요지표채권 전일종가: R54~ (col 4=종목명, 5=종가, 7=전일비, 8=만기, 12=잔액)"""
    count = 0
    # 왼쪽 블록 (예보채, 지역개발, 도시철도, 국주 등)
    sector_map = {}
    current_sector = ''
    for row in range(54, 78):
        s = cell(ws, row, 2)
        if s and str(s).strip() and str(s).strip() != '0':
            current_sector = str(s).strip()
        sector_map[row] = current_sector

    for row in range(54, 78):
        bond_name = cell(ws, row, 4)
        if not bond_name or str(bond_name).strip() == '0' or str(bond_name).strip() == '':
            continue
        bond_name = str(bond_name).strip()
        close_yield = num(cell(ws, row, 5))
        change = num(cell(ws, row, 7))
        mat_date_raw = cell(ws, row, 8)
        mat_date = None
        if isinstance(mat_date_raw, datetime):
            mat_date = mat_date_raw.date()

        outstanding_raw = cell(ws, row, 12)
        outstanding = num(outstanding_raw) if outstanding_raw and str(outstanding_raw).strip() != '-' else None

        sector = sector_map.get(row, '')

        if close_yield is not None:
            cur.execute(
                "INSERT INTO tb_bond_valuation (base_date, bond_alias, sector, standard_name, maturity_date, close_yield, change_bp, m_duration, outstanding) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, bond_name, sector, bond_name, mat_date, close_yield, change, None, outstanding)
            )
            count += 1

    # 오른쪽 블록: 국고채 (col 29=종목, 30=종가, 31=전일비, 32=표준명, 33=만기, 34=MDur, 35=잔액)
    current_tenor = ''
    for row in range(54, 78):
        tenor_marker = cell(ws, row, 28)
        if tenor_marker and str(tenor_marker).strip() in ('2Y', '3Y', '5Y', '10Y', '20Y', '30Y', '50Y'):
            current_tenor = str(tenor_marker).strip()

        bond_name = cell(ws, row, 29)
        if not bond_name or str(bond_name).strip() == '' or str(bond_name).strip() == '0':
            continue
        bond_name = str(bond_name).strip()
        close_yield = num(cell(ws, row, 30))
        change = num(cell(ws, row, 31))
        standard_name = str(cell(ws, row, 32) or bond_name).strip()
        mat_raw = cell(ws, row, 33)
        mat_date = mat_raw.date() if isinstance(mat_raw, datetime) else None
        m_dur = num(cell(ws, row, 34))
        outstanding = num(cell(ws, row, 35))

        if close_yield is not None:
            cur.execute(
                "INSERT INTO tb_bond_valuation (base_date, bond_alias, sector, standard_name, maturity_date, close_yield, change_bp, m_duration, outstanding) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, bond_name, f'국고{current_tenor}', standard_name, mat_date, close_yield, change, m_dur, outstanding)
            )
            count += 1

    print(f"  tb_bond_valuation: {count}건")
    return count


def insert_yield_curve_matrix(ws, cur, base_date):
    """수익률 커브 매트릭스: R79~ (col 4~12 = 3M~20Y 수익률)"""
    tenors = ['3M', '6M', '1Y', '2Y', '3Y', '5Y', '10Y', '20Y']
    tenor_cols = [4, 5, 7, 8, 9, 10, 11, 12]
    count = 0

    for row in range(80, 100):
        name = cell(ws, row, 2)
        if not name or str(name).strip() == '' or str(name).strip() == '0':
            continue
        name_str = str(name).strip()

        # 분류: 국고, 회사채, 통안 등
        if '국고' in name_str or '국주' in name_str:
            sector = '국고'
            credit = ''
        elif '통안' in name_str:
            sector = '통안'
            credit = ''
        elif '산금' in name_str or '수은' in name_str:
            sector = '특수채'
            credit = 'AAA'
        elif 'AA+' in name_str:
            sector = '회사채'
            credit = 'AA+'
        elif 'AA-' in name_str:
            sector = '회사채'
            credit = 'AA-'
        elif 'AA' in name_str:
            sector = '회사채'
            credit = 'AA'
        elif 'A+' in name_str:
            sector = '회사채'
            credit = 'A+'
        elif 'BBB' in name_str:
            sector = '회사채'
            credit = 'BBB'
        else:
            sector = name_str
            credit = ''

        for i, tenor in enumerate(tenors):
            val = num(cell(ws, row, tenor_cols[i]))
            if val is not None:
                cur.execute(
                    "INSERT INTO tb_yield_curve_matrix (base_date, sector, credit_rating, tenor, yield_rate) "
                    "VALUES (%s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                    (base_date, sector, credit, tenor, val)
                )
                count += 1

    print(f"  tb_yield_curve_matrix: {count}건")
    return count


def insert_bond_lending(ws, cur, base_date):
    """채권 대차거래: R37~ (col 26=종목, 28=대차, 30=상환, 32=증감, 33=전전일잔, 34=전일잔)"""
    count = 0
    for row in range(37, 52):
        ticker = cell(ws, row, 26)
        if not ticker or str(ticker).strip() == '' or str(ticker).strip() == '0':
            continue
        ticker_str = str(ticker).strip()
        borrow = num(cell(ws, row, 28))
        repay = num(cell(ws, row, 30))
        net_change = num(cell(ws, row, 32))
        balance = num(cell(ws, row, 34))

        if borrow is not None or balance is not None:
            cur.execute(
                "INSERT INTO tb_bond_lending (base_date, bond_ticker, borrow_amt, repay_amt, net_change, balance) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                (base_date, ticker_str, borrow, repay, net_change, balance)
            )
            count += 1

    print(f"  tb_bond_lending: {count}건")
    return count


def insert_economic_calendar(ws, cur, base_date):
    """경제일정 캘린더 — R100~ 이후 영역 탐색"""
    count = 0
    # 경제일정은 시트 하단에 있을 수 있음 — 넓게 탐색
    for row in range(90, min(ws.max_row + 1, 170)):
        for c in [2, 3, 4, 22, 23]:
            v = cell(ws, row, c)
            if v and '입찰' in str(v) or (v and '발표' in str(v)) or (v and '경매' in str(v)):
                count += 1
                cur.execute(
                    "INSERT INTO tb_economic_calendar (event_date, seq, event_desc) "
                    "VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
                    (base_date, count, str(v).strip())
                )

    # 다음 영업일 일정도 R4:AH4 부근 탐색
    next_date_raw = cell(ws, 4, 34)
    next_date = next_date_raw.date() if isinstance(next_date_raw, datetime) else None
    if next_date:
        # R100~ 이후에 다음일 경제일정이 있을 수 있음
        for row in range(100, min(ws.max_row + 1, 170)):
            for c in range(22, 40):
                v = cell(ws, row, c)
                if v and ('입찰' in str(v) or '발표' in str(v) or '경매' in str(v) or '국고채' in str(v)):
                    count += 1
                    cur.execute(
                        "INSERT INTO tb_economic_calendar (event_date, seq, event_desc) "
                        "VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
                        (next_date, count, str(v).strip())
                    )

    print(f"  tb_economic_calendar: {count}건")
    return count


def process_single_file(filepath, conn):
    """단일 Excel 파일 → DB 삽입. 삽입 건수 반환."""
    print(f"\n파일 로드: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Sub']

    # 기준일 결정: 시트 내부 날짜 우선 (실제 종가 기준일)
    # 파일명은 배포일(당일), 시트 내부 날짜가 실제 데이터 기준일(전 영업일)
    base_date = extract_base_date(ws)
    if not base_date:
        base_date = parse_date_from_filename(filepath)
    if not base_date:
        print("  기준일을 추출할 수 없습니다. 건너뜀.")
        wb.close()
        return 0

    print(f"기준일: {base_date}")

    cur = conn.cursor()
    try:
        total = 0
        total += insert_macro_index(ws, cur, base_date)
        total += insert_domestic_rate(ws, cur, base_date)
        total += insert_ktb_futures(ws, cur, base_date)
        total += insert_primary_market(ws, cur, base_date)
        total += insert_bond_valuation(ws, cur, base_date)
        total += insert_yield_curve_matrix(ws, cur, base_date)
        total += insert_bond_lending(ws, cur, base_date)
        total += insert_economic_calendar(ws, cur, base_date)

        # market_rates 동기화 (overview 페이지 금리 차트용)
        sync_market_rates(cur, base_date)

        conn.commit()
        print(f"  → {total}건 삽입 완료!")
        return total
    except Exception as e:
        conn.rollback()
        print(f"  → 오류 발생: {e}")
        return 0
    finally:
        cur.close()
        wb.close()


def sync_market_rates(cur, base_date):
    """tb_domestic_rate → market_rates 동기화 (overview 금리 차트용)
    국고채 10년 → ktb_10y, UST 10Y → ust_10y"""
    base_date_iso = base_date.isoformat() if isinstance(base_date, date) else str(base_date)
    std_dt = base_date_iso.replace('-', '')  # YYYYMMDD 형식

    # 국고채 10년
    cur.execute(
        "SELECT yield_val FROM tb_domestic_rate WHERE base_date = %s AND rate_type = '국고채' AND maturity = '국고채 10년'",
        (base_date_iso,)
    )
    ktb_row = cur.fetchone()
    ktb_10y = float(ktb_row[0]) if ktb_row else None

    # UST 10Y
    cur.execute(
        "SELECT yield_val FROM tb_domestic_rate WHERE base_date = %s AND rate_type = 'UST' AND maturity = '10Y'",
        (base_date_iso,)
    )
    ust_row = cur.fetchone()
    ust_10y = float(ust_row[0]) if ust_row else None

    if ktb_10y is not None and ust_10y is not None:
        cur.execute("""
            INSERT INTO market_rates (std_dt, ktb_10y, ust_10y)
            VALUES (%s, %s, %s)
            ON CONFLICT (std_dt) DO UPDATE SET ktb_10y = EXCLUDED.ktb_10y, ust_10y = EXCLUDED.ust_10y
        """, (std_dt, ktb_10y, ust_10y))
        print(f"  market_rates 동기화: {std_dt} (KTB10Y={ktb_10y}, UST10Y={ust_10y})")
    else:
        print(f"  market_rates 동기화 건너뜀: 국고10Y={ktb_10y}, UST10Y={ust_10y}")


def get_existing_dates(conn):
    """DB에 이미 존재하는 base_date 집합 반환"""
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT base_date FROM tb_macro_index ORDER BY base_date")
    dates = {row[0] for row in cur.fetchall()}
    cur.close()
    return dates


def extract_base_date_from_file(filepath):
    """Excel 파일의 시트 내부 기준일 추출 (파일을 열어야 함)"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb['Sub']
        for r in range(3, 6):
            for c in range(25, 40):
                v = ws.cell(r, c).value
                if isinstance(v, datetime):
                    wb.close()
                    return v.date()
        wb.close()
    except Exception:
        pass
    return None


def find_daily_files(market_dir='market'):
    """market 폴더에서 *_DAILY.xlsx 파일 목록과 날짜를 반환"""
    files = []
    if not os.path.isdir(market_dir):
        return files
    for fname in sorted(os.listdir(market_dir)):
        if fname.endswith('.xlsx') and ('_DAILY' in fname or '데일리' in fname) and not fname.startswith('~'):
            filepath = os.path.join(market_dir, fname)
            file_date = parse_date_from_filename(filepath)
            if file_date:
                files.append((file_date, filepath))
    return sorted(files, key=lambda x: x[0])


def auto_mode(market_dir='market'):
    """--auto: market 폴더에서 DB에 없는 날짜의 파일을 자동 감지하여 일괄 삽입"""
    print("=" * 60)
    print("AUTO 모드: DB에 없는 날짜 자동 감지 → 일괄 삽입")
    print("=" * 60)

    conn = psycopg2.connect(DB_URL)
    existing = get_existing_dates(conn)
    print(f"DB에 존재하는 날짜: {len(existing)}개 (최근: {max(existing) if existing else 'N/A'})")

    all_files = find_daily_files(market_dir)
    print(f"market 폴더 내 DAILY 파일: {len(all_files)}개")

    # DB에 없는 날짜만 필터 (시트 기준일 기준으로 비교)
    new_files = []
    for d, f in all_files:
        # 시트 내부 기준일 추출하여 비교
        sheet_date = extract_base_date_from_file(f)
        check_date = sheet_date if sheet_date else d
        if check_date not in existing:
            new_files.append((d, f))
    if not new_files:
        print("\n새로 삽입할 파일이 없습니다. DB가 최신 상태입니다.")
        conn.close()
        return

    print(f"\n삽입 대상: {len(new_files)}개 파일")
    for d, f in new_files:
        print(f"  {d} ← {os.path.basename(f)}")

    grand_total = 0
    success = 0
    for d, filepath in new_files:
        count = process_single_file(filepath, conn)
        grand_total += count
        if count > 0:
            success += 1

    print(f"\n{'=' * 60}")
    print(f"완료: {success}/{len(new_files)} 파일, 총 {grand_total}건 삽입")
    print(f"{'=' * 60}")
    conn.close()


def main():
    if len(sys.argv) < 2:
        print("사용법:")
        print("  python3 scripts/bulk_insert_daily.py <엑셀파일>       # 단일 파일 삽입")
        print("  python3 scripts/bulk_insert_daily.py --auto           # DB에 없는 날짜 자동 일괄 삽입")
        print("  python3 scripts/bulk_insert_daily.py --auto market/   # 폴더 지정")
        sys.exit(1)

    if sys.argv[1] == '--auto':
        market_dir = sys.argv[2] if len(sys.argv) > 2 else 'market'
        auto_mode(market_dir)
    else:
        filepath = sys.argv[1]
        if not os.path.exists(filepath):
            print(f"파일을 찾을 수 없습니다: {filepath}")
            sys.exit(1)
        conn = psycopg2.connect(DB_URL)
        process_single_file(filepath, conn)
        conn.close()


if __name__ == '__main__':
    main()
