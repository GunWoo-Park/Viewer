#!/usr/bin/env python3
"""
DataBase 폴더의 TSV 파일을 PostgreSQL에 업로드 (배치 처리)
- breakdownprc, eq_unasp, excpnp: 없는 데이터만 INSERT (ON CONFLICT DO NOTHING)
- strucprdp: UPSERT (obj_cd 기준)

사용법:
  python3 scripts/upload_ficc_tables.py                      # 최신 파일 자동 감지
  python3 scripts/upload_ficc_tables.py --date 20260312      # 날짜(8자리)만 → 해당 날짜 최신 파일
  python3 scripts/upload_ficc_tables.py --date 202603121027  # 전체 타임스탬프도 가능

DataBase 폴더에서 {테이블명}_{타임스탬프}.txt 형식의 최신 파일을 자동 감지합니다.
"""

import sys
import os
import csv
import glob
import psycopg2
from psycopg2.extras import execute_values

DB_URL = "postgresql://neondb_owner:npg_IcbygUXf0Sa4@ep-sweet-smoke-adz6o5mz-pooler.c-2.us-east-1.aws.neon.tech/neondb?sslmode=require"
DB_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'DataBase')

BATCH_SIZE = 500  # 배치 크기


def clean(v):
    """따옴표 제거 및 빈값 처리"""
    if v is None:
        return None
    v = v.strip().strip('"')
    if v == '' or v == 'NULL' or v == '\\N':
        return None
    return v


def num(v):
    """숫자 변환"""
    v = clean(v)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def find_latest_file(table_name, target_ts=None):
    """DataBase 폴더에서 테이블의 최신 파일 찾기
    target_ts: 전체 타임스탬프(202603121027) 또는 날짜만(20260312) 지정 가능
    대소문자 무관, 언더스코어 1~2개 모두 매칭
    """
    # 대소문자/언더스코어 변형 모두 수집
    all_txt = glob.glob(os.path.join(DB_DIR, '*.txt'))
    tname_lower = table_name.lower()
    files = []
    for f in all_txt:
        base = os.path.basename(f).lower()
        # table_name_ 또는 table_name__ 으로 시작하는 파일 매칭
        if base.startswith(f'{tname_lower}_'):
            files.append(f)
    if not files:
        return None

    if target_ts:
        # 정확한 타임스탬프 먼저 시도 (대소문자/언더스코어 변형 포함)
        for f in files:
            base = os.path.basename(f).lower()
            # 테이블명 제거 후 숫자 부분 추출
            rest = base[len(tname_lower):]
            # 언더스코어 제거 후 타임스탬프 비교
            digits = rest.replace('_', '').replace('.txt', '')
            if digits == target_ts:
                return f

        # 날짜(8자리) prefix로 매칭 → 해당 날짜의 최신 파일
        prefix_matches = []
        for f in files:
            base = os.path.basename(f).lower()
            rest = base[len(tname_lower):]
            digits = rest.replace('_', '').replace('.txt', '')
            if digits.startswith(target_ts):
                prefix_matches.append(f)
        if prefix_matches:
            return max(prefix_matches, key=os.path.getmtime)
        return None

    return max(files, key=os.path.getmtime)


def read_tsv(filepath):
    """TSV 파일 읽기 (헤더 + 데이터)"""
    rows = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter='\t')
        headers = [h.strip().strip('"').upper() for h in next(reader)]
        for row in reader:
            if len(row) >= len(headers):
                rows.append(row[:len(headers)])
            elif len(row) > 0:
                padded = row + [None] * (len(headers) - len(row))
                rows.append(padded)
    return headers, rows


def upload_breakdownprc(conn, filepath):
    """breakdownprc: 배치 INSERT (ON CONFLICT DO NOTHING)"""
    print(f"\n[breakdownprc] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    # 기존 키 조회
    cur.execute("SELECT obj_cd || '|' || sp_num || '|' || tp || '|' || std_dt FROM breakdownprc")
    existing = {r[0] for r in cur.fetchall()}
    print(f"  DB 기존: {len(existing)}건")

    # 신규 행만 필터링
    new_rows = []
    skipped = 0
    for row in rows:
        obj_cd = clean(row[0])
        sp_num = clean(row[1])
        tp = clean(row[3])
        std_dt = clean(row[4])
        if not obj_cd or not std_dt:
            continue

        key = f"{obj_cd}|{sp_num}|{tp}|{std_dt}"
        if key in existing:
            skipped += 1
            continue

        asst_lblt = clean(row[2])
        kis_prc = num(row[5])
        kap_prc = num(row[6])
        avg_prc = num(row[7])
        created_at = clean(row[8]) if len(row) > 8 else None
        updated_at = clean(row[9]) if len(row) > 9 else None

        new_rows.append((obj_cd, sp_num, asst_lblt, tp, std_dt,
                         kis_prc or 0, kap_prc or 0, avg_prc or 0,
                         created_at, updated_at))

    # 배치 INSERT
    if new_rows:
        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i + BATCH_SIZE]
            execute_values(cur,
                """INSERT INTO breakdownprc (obj_cd, sp_num, asst_lblt, tp, std_dt, kis_prc, kap_prc, avg_prc, created_at, updated_at)
                   VALUES %s ON CONFLICT (obj_cd, sp_num, tp, std_dt) DO NOTHING""",
                batch, page_size=BATCH_SIZE)
        conn.commit()

    cur.close()
    print(f"  신규 INSERT: {len(new_rows)}건, 스킵(기존): {skipped}건")
    return len(new_rows)


def upload_eq_unasp(conn, filepath):
    """eq_unasp: 배치 INSERT (ON CONFLICT DO NOTHING)"""
    print(f"\n[eq_unasp] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    cur.execute("SELECT std_dt || '|' || unas_id FROM eq_unasp")
    existing = {r[0] for r in cur.fetchall()}
    print(f"  DB 기존: {len(existing)}건")

    new_rows = []
    skipped = 0
    for row in rows:
        std_dt = clean(row[0])
        unas_id = clean(row[1])
        if not std_dt or not unas_id:
            continue

        key = f"{std_dt}|{unas_id}"
        if key in existing:
            skipped += 1
            continue

        reg_dtm = clean(row[2]) if len(row) > 2 else None
        regr_id = clean(row[3]) if len(row) > 3 else None
        mdfy_dtm = clean(row[4]) if len(row) > 4 else None
        mdfyr_id = clean(row[5]) if len(row) > 5 else None
        clprc_val = num(row[6]) if len(row) > 6 else 0
        divd_rt = num(row[7]) if len(row) > 7 else 0
        aply_qnto_val = num(row[8]) if len(row) > 8 else 0
        vltl_val = num(row[9]) if len(row) > 9 else 0

        new_rows.append((std_dt, unas_id, reg_dtm, regr_id, mdfy_dtm, mdfyr_id,
                         clprc_val or 0, divd_rt or 0, aply_qnto_val or 0, vltl_val or 0))

    if new_rows:
        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i + BATCH_SIZE]
            execute_values(cur,
                """INSERT INTO eq_unasp (std_dt, unas_id, reg_dtm, regr_id, mdfy_dtm, mdfyr_id, clprc_val, divd_rt, aply_qnto_val, vltl_val)
                   VALUES %s ON CONFLICT (std_dt, unas_id) DO NOTHING""",
                batch, page_size=BATCH_SIZE)
        conn.commit()

    cur.close()
    print(f"  신규 INSERT: {len(new_rows)}건, 스킵(기존): {skipped}건")
    return len(new_rows)


def upload_excpnp(conn, filepath):
    """excpnp: 배치 INSERT (ON CONFLICT DO NOTHING)"""
    print(f"\n[excpnp] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    cur.execute("SELECT pay_dt || '|' || obj_cd || '|' || tp || '|' || curr || '|' || amt::text FROM excpnp")
    existing = {r[0] for r in cur.fetchall()}
    print(f"  DB 기존: {len(existing)}건")

    new_rows = []
    skipped = 0
    for row in rows:
        pay_dt = clean(row[0])
        fnd_cd = clean(row[1])
        fnd_nm = clean(row[2])
        obj_cd = clean(row[3])
        tp = clean(row[4])
        curr = clean(row[5])
        amt = num(row[6])
        created_at = clean(row[7]) if len(row) > 7 else None
        updated_at = clean(row[8]) if len(row) > 8 else None

        if not pay_dt or not obj_cd:
            continue

        amt_val = amt if amt is not None else 0
        key = f"{pay_dt}|{obj_cd}|{tp}|{curr}|{amt_val}"
        if key in existing:
            skipped += 1
            continue

        new_rows.append((pay_dt, fnd_cd, fnd_nm, obj_cd, tp, curr, amt_val, created_at, updated_at))

    if new_rows:
        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i + BATCH_SIZE]
            execute_values(cur,
                """INSERT INTO excpnp (pay_dt, fnd_cd, fnd_nm, obj_cd, tp, curr, amt, created_at, updated_at)
                   VALUES %s ON CONFLICT (pay_dt, obj_cd, tp, curr, amt) DO NOTHING""",
                batch, page_size=BATCH_SIZE)
        conn.commit()

    cur.close()
    print(f"  신규 INSERT: {len(new_rows)}건, 스킵(기존): {skipped}건")
    return len(new_rows)


def upload_strucprdp(conn, filepath):
    """strucprdp: 배치 UPSERT (obj_cd 기준)"""
    print(f"\n[strucprdp] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    all_rows = []
    for row in rows:
        obj_cd = clean(row[0])
        if not obj_cd:
            continue

        fnd_cd = clean(row[1])
        fnd_nm = clean(row[2])
        cntr_nm = clean(row[3])
        asst_lblt = clean(row[4])
        tp = clean(row[5])
        trd_dt = clean(row[6])
        eff_dt = clean(row[7])
        mat_dt = clean(row[8])
        curr = clean(row[9])
        notn = num(row[10])
        mat_prd = num(row[11])
        call_yn = clean(row[12])
        risk_call_yn = clean(row[13])
        struct_cond = clean(row[14])
        pay_cond = clean(row[15])
        pay_freq = clean(row[16])
        pay_dcf = clean(row[17])
        rcv_cond = clean(row[18])
        rcv_freq = clean(row[19])
        rcv_dcf = clean(row[20])
        note = clean(row[21])
        call_dt = clean(row[22])
        trmntn_dt = clean(row[23])
        type1 = clean(row[24])
        type2 = clean(row[25])
        type3 = clean(row[26])
        type4 = clean(row[27])
        optn_freq = clean(row[28]) if len(row) > 28 else None
        call_notice = clean(row[29]) if len(row) > 29 else None
        frst_call_dt = clean(row[30]) if len(row) > 30 else None
        add_optn = clean(row[31]) if len(row) > 31 else None
        upfrnt = clean(row[32]) if len(row) > 32 else None
        no_val = num(row[33]) if len(row) > 33 else None
        created_at = clean(row[34]) if len(row) > 34 else None
        updated_at = clean(row[35]) if len(row) > 35 else None

        all_rows.append((obj_cd, fnd_cd, fnd_nm, cntr_nm, asst_lblt, tp,
                         trd_dt, eff_dt, mat_dt, curr, notn or 0, mat_prd or 0,
                         call_yn, risk_call_yn, struct_cond, pay_cond, pay_freq, pay_dcf,
                         rcv_cond, rcv_freq, rcv_dcf, note, call_dt, trmntn_dt,
                         type1, type2, type3, type4, optn_freq, call_notice,
                         frst_call_dt, add_optn, upfrnt,
                         int(no_val) if no_val else None, created_at, updated_at))

    if all_rows:
        for i in range(0, len(all_rows), BATCH_SIZE):
            batch = all_rows[i:i + BATCH_SIZE]
            execute_values(cur,
                """INSERT INTO strucprdp (obj_cd, fnd_cd, fnd_nm, cntr_nm, asst_lblt, tp,
                    trd_dt, eff_dt, mat_dt, curr, notn, mat_prd, call_yn, risk_call_yn,
                    struct_cond, pay_cond, pay_freq, pay_dcf, rcv_cond, rcv_freq, rcv_dcf,
                    note, call_dt, trmntn_dt, type1, type2, type3, type4,
                    optn_freq, call_notice, frst_call_dt, add_optn, upfrnt, no, created_at, updated_at)
                   VALUES %s
                   ON CONFLICT (obj_cd) DO UPDATE SET
                    fnd_cd=EXCLUDED.fnd_cd, fnd_nm=EXCLUDED.fnd_nm, cntr_nm=EXCLUDED.cntr_nm,
                    asst_lblt=EXCLUDED.asst_lblt, tp=EXCLUDED.tp, trd_dt=EXCLUDED.trd_dt,
                    eff_dt=EXCLUDED.eff_dt, mat_dt=EXCLUDED.mat_dt, curr=EXCLUDED.curr,
                    notn=EXCLUDED.notn, mat_prd=EXCLUDED.mat_prd, call_yn=EXCLUDED.call_yn,
                    risk_call_yn=EXCLUDED.risk_call_yn, struct_cond=EXCLUDED.struct_cond,
                    pay_cond=EXCLUDED.pay_cond, pay_freq=EXCLUDED.pay_freq, pay_dcf=EXCLUDED.pay_dcf,
                    rcv_cond=EXCLUDED.rcv_cond, rcv_freq=EXCLUDED.rcv_freq, rcv_dcf=EXCLUDED.rcv_dcf,
                    note=EXCLUDED.note, call_dt=EXCLUDED.call_dt, trmntn_dt=EXCLUDED.trmntn_dt,
                    type1=EXCLUDED.type1, type2=EXCLUDED.type2, type3=EXCLUDED.type3, type4=EXCLUDED.type4,
                    optn_freq=EXCLUDED.optn_freq, call_notice=EXCLUDED.call_notice,
                    frst_call_dt=EXCLUDED.frst_call_dt, add_optn=EXCLUDED.add_optn, upfrnt=EXCLUDED.upfrnt,
                    no=EXCLUDED.no, updated_at=EXCLUDED.updated_at""",
                batch, page_size=BATCH_SIZE)
        conn.commit()

    cur.close()
    print(f"  UPSERT: {len(all_rows)}건")
    return len(all_rows)


def upload_pnlrtp(conn, filepath):
    """pnlrtp: 배치 INSERT (ON CONFLICT DO NOTHING)"""
    print(f"\n[pnlrtp] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    # 기존 키 조회
    cur.execute("SELECT std_dt || '|' || intl_ytm_curv_cd || '|' || curv_tp_cd || '|' || tnr_cd FROM pnlrtp")
    existing = {r[0] for r in cur.fetchall()}
    print(f"  DB 기존: {len(existing)}건")

    new_rows = []
    skipped = 0
    for row in rows:
        std_dt = clean(row[0])
        intl_ytm_curv_cd = clean(row[1])
        curv_tp_cd = clean(row[2])
        tnr_cd = clean(row[3])
        if not std_dt or not tnr_cd:
            continue

        key = f"{std_dt}|{intl_ytm_curv_cd}|{curv_tp_cd}|{tnr_cd}"
        if key in existing:
            skipped += 1
            continue

        reg_dtm = clean(row[4])
        regr_id = clean(row[5])
        mdfy_dtm = clean(row[6])
        mdfyr_id = clean(row[7])
        nt_tnr_dlt = num(row[8])
        str_tnr_dlt = num(row[9])
        str_tnr_delta_chg = num(row[10])
        hdg_tnr_dlt = num(row[11])
        hdg_tnr_dlt_chg = num(row[12])
        sprdvl_sprd = num(row[13])
        sprdvl_sprd_yr = num(row[14])
        sprdvl_dlt = num(row[15])
        sprdvl_crry_d = num(row[16])
        crry = num(row[17])
        rll = num(row[18])

        new_rows.append((
            std_dt, intl_ytm_curv_cd, curv_tp_cd, tnr_cd,
            reg_dtm, regr_id, mdfy_dtm, mdfyr_id,
            nt_tnr_dlt, str_tnr_dlt, str_tnr_delta_chg,
            hdg_tnr_dlt, hdg_tnr_dlt_chg,
            sprdvl_sprd, sprdvl_sprd_yr, sprdvl_dlt, sprdvl_crry_d,
            crry, rll,
        ))

    if new_rows:
        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i + BATCH_SIZE]
            execute_values(cur,
                """INSERT INTO pnlrtp (
                    std_dt, intl_ytm_curv_cd, curv_tp_cd, tnr_cd,
                    reg_dtm, regr_id, mdfy_dtm, mdfyr_id,
                    nt_tnr_dlt, str_tnr_dlt, str_tnr_delta_chg,
                    hdg_tnr_dlt, hdg_tnr_dlt_chg,
                    sprdvl_sprd, sprdvl_sprd_yr, sprdvl_dlt, sprdvl_crry_d,
                    crry, rll
                ) VALUES %s ON CONFLICT (std_dt, intl_ytm_curv_cd, curv_tp_cd, tnr_cd) DO NOTHING""",
                batch, page_size=BATCH_SIZE)
        conn.commit()

    cur.close()
    print(f"  신규 INSERT: {len(new_rows)}건, 스킵(기존): {skipped}건")
    return len(new_rows)


def upload_market_rates(conn, target_ts=None):
    """시장 Excel 파일에서 KTB 10Y, UST 10Y 금리 추출 → market_rates UPSERT
    - market/ 폴더의 YYMMDD_DAILY.xlsx 또는 시장/ 폴더의 시장_YYYYMMDD.xlsx 지원
    - 엑셀 T일 시초가 = T-1일 종가 → pnlrtp의 T-1일(std_dt)에 매핑
    """
    import openpyxl
    import re

    project_root = os.path.dirname(DB_DIR)

    # pnlrtp 영업일 목록 (T-1 매핑용)
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT std_dt FROM pnlrtp ORDER BY std_dt")
    biz_days = sorted([r[0] for r in cur.fetchall()])

    if not biz_days:
        print("\n[market_rates] pnlrtp에 데이터 없음, 스킵")
        cur.close()
        return 0

    # 시장 Excel 파일 수집: market/ 폴더 (YYMMDD_DAILY.xlsx) + 시장/ 폴더 (시장_YYYYMMDD.xlsx)
    market_entries = []  # (filepath, file_date_YYYYMMDD)

    # 1) market/ 폴더: YYMMDD_DAILY.xlsx → 20YYMMDD
    market_dir1 = os.path.join(project_root, 'market')
    if os.path.exists(market_dir1):
        for f in glob.glob(os.path.join(market_dir1, '*_DAILY.xlsx')):
            fname = os.path.basename(f)
            m = re.match(r'^(\d{6})_DAILY', fname)
            if m:
                market_entries.append((f, '20' + m.group(1)))

    # 2) 시장/ 폴더: 시장_YYYYMMDD.xlsx
    market_dir2 = os.path.join(project_root, '시장')
    if os.path.exists(market_dir2):
        for f in glob.glob(os.path.join(market_dir2, '시장_*.xlsx')):
            fname = os.path.basename(f)
            parts = fname.replace('.xlsx', '').split('_')
            if len(parts) >= 2 and len(parts[1]) >= 8:
                market_entries.append((f, parts[1][:8]))

    if not market_entries:
        print(f"\n[market_rates] 시장 Excel 파일 없음 (market/, 시장/ 모두 확인)")
        cur.close()
        return 0

    # target_ts 필터 (날짜 지정 시 해당 날짜 파일만)
    if target_ts:
        date_prefix = target_ts[:8]
        market_entries = [(f, d) for f, d in market_entries if d.startswith(date_prefix)]

    print(f"\n[market_rates] {len(market_entries)}개 시장 파일 처리")

    # T-day → T-1 영업일 매핑 (엑셀 T일 시초가 = T-1일 종가)
    def get_prev_biz_day(dt_str):
        """dt_str보다 작은 가장 큰 영업일"""
        for i in range(len(biz_days) - 1, -1, -1):
            if biz_days[i] < dt_str:
                return biz_days[i]
        return None

    inserted = 0
    for filepath, file_date in sorted(market_entries, key=lambda x: x[1]):
        fname = os.path.basename(filepath)
        # T-1 영업일 매핑: 엑셀 T일 시초가 → T-1일 종가
        mapped_date = get_prev_biz_day(file_date)
        if not mapped_date:
            continue

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active

            # KTB 10Y: X17 (열24, 행17)
            ktb_val = ws.cell(row=17, column=24).value
            # UST 10Y: J11 (열10, 행11)
            ust_val = ws.cell(row=11, column=10).value
            wb.close()

            ktb_10y = float(ktb_val) if ktb_val is not None else None
            ust_10y = float(ust_val) if ust_val is not None else None

            if ktb_10y is None and ust_10y is None:
                continue

            cur.execute(
                """INSERT INTO market_rates (std_dt, ktb_10y, ust_10y)
                   VALUES (%s, %s, %s)
                   ON CONFLICT (std_dt) DO UPDATE SET
                     ktb_10y = EXCLUDED.ktb_10y,
                     ust_10y = EXCLUDED.ust_10y""",
                (mapped_date, ktb_10y, ust_10y)
            )
            inserted += 1
        except Exception as e:
            print(f"  {fname} 오류: {e}")
            continue

    conn.commit()
    cur.close()
    print(f"  UPSERT: {inserted}건")
    return inserted


def upload_swap_prc(conn, filepath):
    """swap_prc: 배치 INSERT (std_dt + obj_cd 기준 중복 스킵)"""
    print(f"\n[swap_prc] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    # 기존 키 조회 (std_dt + obj_cd)
    cur.execute("SELECT std_dt || '|' || obj_cd FROM swap_prc")
    existing = {r[0] for r in cur.fetchall()}
    print(f"  DB 기존: {len(existing)}건")

    new_rows = []
    skipped = 0
    for row in rows:
        no_val = num(row[0])
        std_dt = clean(row[1])
        fnd_cd = clean(row[2])
        fnd_nm = clean(row[3])
        obj_cd = clean(row[4])
        if not std_dt or not obj_cd:
            continue

        key = f"{std_dt}|{obj_cd}"
        if key in existing:
            skipped += 1
            continue

        obj_nm = clean(row[5])
        prd_dvsn = clean(row[6])
        prd_grp = clean(row[7])
        asst_lvl = clean(row[8])
        trd_tp = clean(row[9])
        cntrprt = clean(row[10])
        pstn = clean(row[11])
        fxd_rt = num(row[12])
        init_prncpl_ex = clean(row[13])
        lst_prncpl_ex = clean(row[14])
        crrnt2 = clean(row[15])
        ntnl2 = num(row[16])
        crrnt1 = clean(row[17])
        ntnl1 = num(row[18])
        kis_prc = num(row[19])
        kap_prc = num(row[20])
        nic_prc = num(row[21])
        fnp_prc = num(row[22])
        lstyr_prc = num(row[23])
        avg_prc = num(row[24])
        buy_pnl = num(row[25])
        buy_pnl_accnt = num(row[26])
        ystrdy_npv = num(row[27])
        tdy_npv = num(row[28])
        calc_blnc = num(row[29])
        swp_asst_lblt = num(row[30])
        swp_asst_lblt_accnt = num(row[31])
        prmm_asst_lblt = num(row[32])
        trd_dt = clean(row[33])
        eff_dt = clean(row[34])
        adjst_mtrty = clean(row[35])
        inpt_mtrty = clean(row[36])
        mar = num(row[37])
        rcv_ytm = num(row[38])
        rcv_dur = num(row[39])
        pay_ytm = num(row[40])
        pay_dur = num(row[41])
        swap_dur = num(row[42])
        prc_diff = clean(row[43])
        avg_prc_usd = num(row[44])
        nd_yn = clean(row[45]) if len(row) > 45 else None
        erly_exr_yn = clean(row[46]) if len(row) > 46 else None
        erly_exr_dt = clean(row[47]) if len(row) > 47 else None

        new_rows.append((
            int(no_val) if no_val else None, std_dt, fnd_cd, fnd_nm, obj_cd, obj_nm,
            prd_dvsn, prd_grp, asst_lvl, trd_tp, cntrprt, pstn,
            fxd_rt, init_prncpl_ex, lst_prncpl_ex,
            crrnt2, ntnl2, crrnt1, ntnl1,
            kis_prc, kap_prc, nic_prc, fnp_prc, lstyr_prc, avg_prc,
            buy_pnl, buy_pnl_accnt, ystrdy_npv, tdy_npv, calc_blnc,
            swp_asst_lblt, swp_asst_lblt_accnt, prmm_asst_lblt,
            trd_dt, eff_dt, adjst_mtrty, inpt_mtrty,
            mar, rcv_ytm, rcv_dur, pay_ytm, pay_dur, swap_dur,
            prc_diff, avg_prc_usd, nd_yn, erly_exr_yn, erly_exr_dt,
        ))

    if new_rows:
        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i + BATCH_SIZE]
            execute_values(cur,
                """INSERT INTO swap_prc (
                    no, std_dt, fnd_cd, fnd_nm, obj_cd, obj_nm,
                    prd_dvsn, prd_grp, asst_lvl, trd_tp, cntrprt, pstn,
                    fxd_rt, init_prncpl_ex, lst_prncpl_ex,
                    crrnt2, ntnl2, crrnt1, ntnl1,
                    kis_prc, kap_prc, nic_prc, fnp_prc, lstyr_prc, avg_prc,
                    buy_pnl, buy_pnl_accnt, ystrdy_npv, tdy_npv, calc_blnc,
                    swp_asst_lblt, swp_asst_lblt_accnt, prmm_asst_lblt,
                    trd_dt, eff_dt, adjst_mtrty, inpt_mtrty,
                    mar, rcv_ytm, rcv_dur, pay_ytm, pay_dur, swap_dur,
                    prc_diff, avg_prc_usd, nd_yn, erly_exr_yn, erly_exr_dt
                ) VALUES %s""",
                batch, page_size=BATCH_SIZE)
        conn.commit()

    cur.close()
    print(f"  신규 INSERT: {len(new_rows)}건, 스킵(기존): {skipped}건")
    return len(new_rows)


def upload_strucfe_accint(conn, filepath):
    """strucfe_accint: INSERT (기존 데이터 보존, 중복 키는 스킵)
    파일명 패턴: struc_accint_*.txt (테이블명은 strucfe_accint)
    파일 컬럼이 DB보다 많으므로 필요한 컬럼만 매핑
    """
    print(f"\n[strucfe_accint] {os.path.basename(filepath)}")
    headers, rows = read_tsv(filepath)
    cur = conn.cursor()

    # 헤더 인덱스 매핑
    hmap = {h: i for i, h in enumerate(headers)}

    all_rows = []
    for row in rows:
        obj_cd = clean(row[hmap.get('OBJ_CD', 0)])
        if not obj_cd:
            continue
        nx_cd = clean(row[hmap.get('NX_CD', 1)])
        eval_mdul_cd = clean(row[hmap.get('EVAL_MDUL_CD', 2)])
        leg_tp = clean(row[hmap.get('LEG_TP', 3)])
        std_dt = clean(row[hmap.get('STD_DT', 7)])
        eval_dt = clean(row[hmap.get('EVAL_DT', 8)])
        cf_num_val = num(row[hmap.get('CF_NUM', 9)])
        reg_dtm = clean(row[hmap.get('REG_DTM', 10)])
        regr_id = clean(row[hmap.get('REGR_ID', 11)])
        mdfy_dtm = clean(row[hmap.get('MDFY_DTM', 12)])
        mdfyr_id = clean(row[hmap.get('MDFYR_ID', 13)])
        str_dt = clean(row[hmap.get('STR_DT', 14)])
        end_dt = clean(row[hmap.get('END_DT', 15)])
        pay_dt = clean(row[hmap.get('PAY_DT', 16)])
        rate = num(row[hmap.get('RATE', 20)])
        cpn = num(row[hmap.get('CPN', 21)])
        accint = num(row[hmap.get('ACCINT', 22)])

        if not std_dt or not leg_tp:
            continue

        all_rows.append((
            obj_cd, nx_cd, eval_mdul_cd, leg_tp, std_dt, eval_dt,
            int(cf_num_val) if cf_num_val else None,
            reg_dtm, regr_id, mdfy_dtm, mdfyr_id,
            str_dt, end_dt, pay_dt,
            rate or 0, cpn or 0, accint or 0,
        ))

    inserted = 0
    skipped = 0
    if all_rows:
        for i in range(0, len(all_rows), BATCH_SIZE):
            batch = all_rows[i:i + BATCH_SIZE]
            cur.execute("SAVEPOINT batch_sp")
            try:
                execute_values(cur,
                    """INSERT INTO strucfe_accint (
                        obj_cd, nx_cd, eval_mdul_cd, leg_tp, std_dt, eval_dt,
                        cf_num, reg_dtm, regr_id, mdfy_dtm, mdfyr_id,
                        str_dt, end_dt, pay_dt, rate, cpn, accint
                    ) VALUES %s
                    ON CONFLICT (obj_cd, leg_tp, std_dt) DO NOTHING""",
                    batch, page_size=BATCH_SIZE)
                inserted += cur.rowcount
                skipped += len(batch) - cur.rowcount
            except Exception:
                cur.execute("ROLLBACK TO SAVEPOINT batch_sp")
                raise
        conn.commit()

    cur.close()
    print(f"  신규 INSERT: {inserted}건, 스킵(기존): {skipped}건")
    return inserted


def find_struc_accint_file(target_ts=None):
    """struc_accint 파일 찾기 (파일명: struc_accint_*.txt, DB 테이블: strucfe_accint)"""
    all_txt = glob.glob(os.path.join(DB_DIR, 'struc_accint*.txt'))
    if not all_txt:
        return None
    if target_ts:
        for f in all_txt:
            base = os.path.basename(f).lower()
            rest = base.replace('struc_accint', '')
            digits = rest.replace('_', '').replace('.txt', '')
            if digits == target_ts or digits.startswith(target_ts):
                return f
    return max(all_txt, key=os.path.getmtime)


def find_swap_prc_file(target_ts=None):
    """6751_swap_prc 파일 찾기 (파일명 패턴: 6751_swap_prc__*.txt 또는 6751_swap_prc_*.txt)"""
    all_txt = glob.glob(os.path.join(DB_DIR, '6751_swap_prc*.txt'))
    if not all_txt:
        return None
    if target_ts:
        for f in all_txt:
            base = os.path.basename(f).lower()
            rest = base.replace('6751_swap_prc', '')
            digits = rest.replace('_', '').replace('.txt', '')
            if digits == target_ts or digits.startswith(target_ts):
                return f
    return max(all_txt, key=os.path.getmtime)


def main():
    target_ts = None
    if len(sys.argv) > 1 and sys.argv[1] == '--date' and len(sys.argv) > 2:
        target_ts = sys.argv[2]

    print("=" * 60)
    print("FICC 테이블 업로드")
    print(f"DataBase 폴더: {DB_DIR}")
    print("=" * 60)

    # 각 테이블의 최신 파일 찾기
    tables = {
        'breakdownprc': upload_breakdownprc,
        'eq_unasp': upload_eq_unasp,
        'excpnp': upload_excpnp,
        'strucprdp': upload_strucprdp,
        'pnlrtp': upload_pnlrtp,
    }

    files_found = {}
    for table_name in tables:
        f = find_latest_file(table_name, target_ts)
        if f:
            files_found[table_name] = f
            print(f"  {table_name}: {os.path.basename(f)}")
        else:
            print(f"  {table_name}: 파일 없음!")

    # swap_prc (6751_swap_prc 패턴)
    swap_prc_file = find_swap_prc_file(target_ts)
    if swap_prc_file:
        files_found['swap_prc'] = swap_prc_file
        print(f"  swap_prc: {os.path.basename(swap_prc_file)}")
    else:
        print(f"  swap_prc: 파일 없음!")

    # strucfe_accint (struc_accint_*.txt 패턴)
    accint_file = find_struc_accint_file(target_ts)
    if accint_file:
        files_found['strucfe_accint'] = accint_file
        print(f"  strucfe_accint: {os.path.basename(accint_file)}")
    else:
        print(f"  strucfe_accint: 파일 없음!")

    if not files_found:
        print("\n업로드할 파일이 없습니다.")
        return

    conn = psycopg2.connect(DB_URL)
    total = 0
    try:
        for table_name, filepath in files_found.items():
            if table_name == 'swap_prc':
                total += upload_swap_prc(conn, filepath)
                continue
            if table_name == 'strucfe_accint':
                total += upload_strucfe_accint(conn, filepath)
                continue
            fn = tables[table_name]
            total += fn(conn, filepath)

        # market_rates: 시장 Excel → 금리 데이터 업로드
        total += upload_market_rates(conn, target_ts)
    except Exception as e:
        conn.rollback()
        print(f"\n오류 발생: {e}")
        raise
    finally:
        conn.close()

    # call_yn 자동 감지: 발행일 이후 MTM=0이면 call_yn='Y'로 업데이트
    print(f"\n--- call_yn 자동 감지 ---")
    conn2 = psycopg2.connect(DB_URL)
    try:
        cur = conn2.cursor()
        cur.execute("""
            WITH zero_mtm AS (
                SELECT s.obj_cd, s.eff_dt, b.std_dt, SUM(b.avg_prc) AS total_mtm
                FROM strucprdp s
                JOIN breakdownprc b ON b.obj_cd = s.obj_cd
                WHERE s.call_yn = 'N' AND b.std_dt > s.eff_dt
                GROUP BY s.obj_cd, s.eff_dt, b.std_dt
                HAVING ABS(SUM(b.avg_prc)) < 0.01
            ),
            first_zero AS (
                SELECT obj_cd, MIN(std_dt) AS call_dt
                FROM zero_mtm GROUP BY obj_cd
            )
            SELECT obj_cd, call_dt FROM first_zero ORDER BY obj_cd
        """)
        call_targets = cur.fetchall()
        if call_targets:
            for obj_cd, call_dt in call_targets:
                cur.execute(
                    "UPDATE strucprdp SET call_yn = 'Y', call_dt = %s, updated_at = NOW() WHERE obj_cd = %s AND call_yn = 'N'",
                    (call_dt, obj_cd)
                )
                print(f"  {obj_cd}: call_yn → Y (call_dt={call_dt})")
            conn2.commit()
            print(f"  → {len(call_targets)}건 업데이트")
        else:
            print("  → 감지된 신규 Call 종목 없음")
    except Exception as e:
        conn2.rollback()
        print(f"  call_yn 감지 오류: {e}")
    finally:
        conn2.close()

    print(f"\n{'=' * 60}")
    print(f"완료: 총 {total}건 처리")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
